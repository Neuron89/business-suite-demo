import json
import os
from datetime import date, datetime
from functools import wraps

from dotenv import load_dotenv

load_dotenv()  # picks up PORTAL_SSO_SECRET (and any other vars) from .env

from io import BytesIO

from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_wtf.csrf import CSRFProtect

from config import Config
from models import (
    AuditLog,
    Item,
    Lot,
    SampleSet,
    Supervisor,
    SystemRequest,
    TemplateParameter,
    TestParameter,
    TestResult,
    TestTemplate,
    User,
    db,
)

# ---------------------------------------------------------------------------
# App factory
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
csrf = CSRFProtect(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message_category = "info"


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def log_audit(action, table_name=None, record_id=None, details=None):
    """Write an entry to the audit log."""
    entry = AuditLog(
        user_id=current_user.id,
        action=action,
        table_name=table_name,
        record_id=record_id,
        details=details,
        ip_address=request.remote_addr,
    )
    db.session.add(entry)
    db.session.commit()


def admin_required(f):
    """Decorator that restricts a view to admin users."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            flash("You do not have permission to access this page.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated


def lab_owner_or_admin_required(f):
    """Decorator that restricts a view to lab_owner or admin users."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.can_manage_items:
            flash("You do not have permission to access this page.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password) and user.is_active_user:
            login_user(user)
            log_audit("login")
            flash(f"Welcome back, {user.full_name}!", "success")
            next_page = request.args.get("next")
            return redirect(next_page or url_for("dashboard"))
        else:
            flash("Invalid username or password.", "error")

    users = User.query.filter_by(is_active_user=True).order_by(User.full_name).all()
    return render_template("login.html", users=users)


@app.route("/logout")
@login_required
def logout():
    log_audit("logout")
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/sso")
def portal_sso():
    """Portal single-sign-on landing.

    The portal redirects browsers here with `?token=<jwt>&next=<path>` after
    the user has authenticated upstream. We verify the JWT against the shared
    PORTAL_SSO_SECRET, find or create a User row keyed by email, and call
    Flask-Login's `login_user()` so they bypass the kiosk login screen.

    The kiosk login flow at `/login` is intentionally untouched — touchscreen
    operators still type a password to switch users on the QC Lab terminal.
    """
    try:
        import jwt as _jwt
    except ImportError:
        return ("PyJWT not installed.", 503)
    secret = os.environ.get("PORTAL_SSO_SECRET")
    if not secret:
        return ("SSO not configured.", 503)
    token = request.args.get("token") or request.args.get("ptoken")
    next_url = request.args.get("next") or url_for("dashboard")
    if not next_url.startswith("/"):
        next_url = url_for("dashboard")
    if not token:
        return ("Missing SSO token.", 400)
    try:
        claims = _jwt.decode(
            token,
            secret,
            algorithms=["HS256"],
            audience="qc",
            issuer="acme-portal",
        )
    except _jwt.PyJWTError:
        return ("Invalid or expired SSO token.", 401)
    email = str(claims.get("email", "")).strip().lower()
    if not email:
        return ("SSO token missing email.", 400)
    full_name = str(claims.get("full_name") or email)
    is_admin = claims.get("portal_role") == "admin"
    role = "admin" if is_admin else "operator"

    user = User.query.filter_by(username=email).first()
    if not user:
        user = User(
            username=email,
            full_name=full_name,
            role=role,
            is_active_user=True,
            can_approve_edits=is_admin,
        )
        # password_hash is NOT NULL but a portal-SSO user never types a
        # password; set an unusable placeholder.
        user.set_password("!sso-no-password")
        db.session.add(user)
        db.session.commit()
    else:
        changed = False
        if not user.is_active_user:
            user.is_active_user = True
            changed = True
        if is_admin and user.role != "admin":
            user.role = "admin"
            changed = True
        if changed:
            db.session.commit()

    login_user(user)
    log_audit("login_sso")
    return redirect(next_url)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------


@app.route("/")
@login_required
def dashboard():
    stats = {
        "total_lots": Lot.query.count(),
        "pending_lots": Lot.query.filter_by(status="pending").count(),
        "completed_lots": Lot.query.filter_by(status="completed").count(),
        "flagged_lots": Lot.query.filter(
            (Lot.status == "flagged") | (Lot.is_duplicate == True)  # noqa: E712
        ).count(),
    }
    recent_lots = Lot.query.order_by(Lot.created_at.desc()).limit(10).all()
    recent_activity = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(15).all()

    return render_template(
        "dashboard.html",
        stats=stats,
        recent_lots=recent_lots,
        recent_activity=recent_activity,
    )


# ---------------------------------------------------------------------------
# Kiosk – barcode scan / new lot
# ---------------------------------------------------------------------------


@app.route("/kiosk", methods=["GET", "POST"])
@login_required
def kiosk():
    items = Item.query.order_by(Item.item_number).all()

    if request.method == "POST":
        barcode = request.form.get("barcode", "").strip()
        item_id = request.form.get("item_id")
        date_received = request.form.get("date_received")
        machine = request.form.get("machine", "").strip()

        if not barcode or not item_id:
            flash("Barcode/Lot number and Item are required.", "error")
            return render_template("kiosk.html", items=items, today=date.today().isoformat())

        # Check for duplicates
        existing = Lot.query.filter_by(lot_number=barcode).first()
        is_dup = existing is not None

        new_lot = Lot(
            lot_number=barcode,
            barcode=barcode,
            item_id=int(item_id),
            date_received=datetime.strptime(date_received, "%Y-%m-%d").date() if date_received else None,
            machine=machine or None,
            status="in_progress",
            is_duplicate=is_dup,
            duplicate_of_id=existing.id if is_dup else None,
            created_by=current_user.id,
        )
        db.session.add(new_lot)
        db.session.commit()

        detail = f"Lot '{barcode}' created for item {item_id}"
        if is_dup:
            detail += f" [DUPLICATE of Lot #{existing.id}]"
        log_audit("create", "lots", new_lot.id, detail)

        if is_dup:
            flash(
                f"Lot '{barcode}' created but flagged as DUPLICATE (original: Lot #{existing.id}).",
                "warning",
            )
        else:
            flash(f"Lot '{barcode}' created successfully!", "success")

        return redirect(url_for("test_entry", lot_id=new_lot.id))

    return render_template("kiosk.html", items=items, today=date.today().isoformat())


# ---------------------------------------------------------------------------
# API – duplicate check (AJAX)
# ---------------------------------------------------------------------------


@app.route("/api/check-duplicate")
@login_required
def check_duplicate():
    lot_number = request.args.get("lot_number", "").strip()
    if not lot_number:
        return jsonify({"is_duplicate": False})

    existing = Lot.query.filter_by(lot_number=lot_number).first()
    if existing:
        return jsonify(
            {
                "is_duplicate": True,
                "lot_id": existing.id,
                "item_number": existing.item.item_number,
                "status": existing.status,
                "created_at": existing.created_at.strftime("%m/%d/%Y %I:%M %p"),
            }
        )
    return jsonify({"is_duplicate": False})


# ---------------------------------------------------------------------------
# Lots list
# ---------------------------------------------------------------------------


@app.route("/lots")
@login_required
def lots_list():
    query = Lot.query
    items = Item.query.order_by(Item.item_number).all()

    # Filters
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    item_id = request.args.get("item_id", "").strip()
    dups_only = request.args.get("duplicates_only")

    if search:
        query = query.filter(
            (Lot.lot_number.ilike(f"%{search}%")) | (Lot.barcode.ilike(f"%{search}%"))
        )
    if status:
        query = query.filter_by(status=status)
    if item_id:
        query = query.filter_by(item_id=int(item_id))
    if dups_only:
        query = query.filter_by(is_duplicate=True)

    # Build query string for pagination links
    qs_parts = []
    for k in ("search", "status", "item_id", "duplicates_only"):
        v = request.args.get(k, "")
        if v:
            qs_parts.append(f"{k}={v}")
    query_string = "&".join(qs_parts)

    page = request.args.get("page", 1, type=int)
    pagination = query.order_by(Lot.created_at.desc()).paginate(
        page=page, per_page=25, error_out=False
    )

    return render_template(
        "lots_list.html",
        lots=pagination.items,
        pagination=pagination,
        items=items,
        query_string=query_string,
    )


# ---------------------------------------------------------------------------
# Export lots to Excel
# ---------------------------------------------------------------------------


@app.route("/lots/export")
@login_required
def export_lots():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    export_type = request.args.get("type", "full")  # full, date_range, failed
    query = Lot.query

    # --- Build query based on export type ---
    if export_type == "failed":
        # Get lots that have at least one failing test result
        failing_lot_ids = (
            db.session.query(TestResult.lot_id)
            .filter(TestResult.is_passing == False)  # noqa: E712
            .distinct()
            .subquery()
        )
        query = query.filter(Lot.id.in_(db.session.query(failing_lot_ids)))
        filename_label = "Failed"

    elif export_type == "date_range":
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        if date_from:
            query = query.filter(
                Lot.created_at >= datetime.strptime(date_from, "%Y-%m-%d")
            )
        if date_to:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59
            )
            query = query.filter(Lot.created_at <= dt_to)
        filename_label = "DateRange"

    else:
        filename_label = "Full"

    lots = query.order_by(Lot.created_at.desc()).all()

    wb = Workbook()

    # --- Sheet 1: Lots Summary ---
    ws = wb.active
    ws.title = "Lots"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

    headers = [
        "Lot #", "Item #", "Item Name", "Date Received", "Status",
        "Result", "Duplicate", "Machine", "Created By", "Created At",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, lot in enumerate(lots, 2):
        passing = lot.is_passing
        if passing is None:
            result = "No Results"
        elif passing:
            result = "PASS"
        else:
            result = "FAIL"

        ws.cell(row=row_idx, column=1, value=lot.lot_number)
        ws.cell(row=row_idx, column=2, value=lot.item.item_number)
        ws.cell(row=row_idx, column=3, value=lot.item.item_name)
        ws.cell(row=row_idx, column=4, value=lot.date_received.strftime("%m/%d/%Y") if lot.date_received else "")
        ws.cell(row=row_idx, column=5, value=lot.status.replace("_", " ").title())
        ws.cell(row=row_idx, column=6, value=result)
        ws.cell(row=row_idx, column=7, value="Yes" if lot.is_duplicate else "No")
        ws.cell(row=row_idx, column=8, value=lot.machine or "")
        ws.cell(row=row_idx, column=9, value=lot.creator.full_name)
        ws.cell(row=row_idx, column=10, value=lot.created_at.strftime("%m/%d/%Y %I:%M %p"))

    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    # --- Sheet 2: Test Results ---
    ws2 = wb.create_sheet("Test Results")

    headers2 = [
        "Lot #", "Item #", "Sample Set", "Parameter", "Value", "Unit",
        "Min Spec", "Max Spec", "Pass/Fail", "Entered By", "Date",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row_idx = 2
    for lot in lots:
        for tr in lot.test_results.order_by(TestResult.entered_at).all():
            if tr.is_passing is None:
                pf = ""
            elif tr.is_passing:
                pf = "PASS"
            else:
                pf = "FAIL"

            ws2.cell(row=row_idx, column=1, value=lot.lot_number)
            ws2.cell(row=row_idx, column=2, value=lot.item.item_number)
            ws2.cell(row=row_idx, column=3, value=tr.sample_set.label if tr.sample_set else "A")
            ws2.cell(row=row_idx, column=4, value=tr.parameter_name)
            ws2.cell(row=row_idx, column=5, value=tr.numeric_value if tr.numeric_value is not None else tr.value)
            ws2.cell(row=row_idx, column=6, value=tr.unit or "")
            ws2.cell(row=row_idx, column=7, value=tr.min_spec if tr.min_spec is not None else "")
            ws2.cell(row=row_idx, column=8, value=tr.max_spec if tr.max_spec is not None else "")
            ws2.cell(row=row_idx, column=9, value=pf)
            ws2.cell(row=row_idx, column=10, value=tr.entered_by_user.full_name)
            ws2.cell(row=row_idx, column=11, value=tr.entered_at.strftime("%m/%d/%Y %I:%M %p") if tr.entered_at else "")
            row_idx += 1

    for col in ws2.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, min(len(val), 40))
        ws2.column_dimensions[col[0].column_letter].width = max_len + 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"QC_{filename_label}_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# Lot detail
# ---------------------------------------------------------------------------


@app.route("/lots/<int:lot_id>")
@login_required
def lot_detail(lot_id):
    lot = db.session.get(Lot, lot_id)
    if not lot:
        flash("Lot not found.", "error")
        return redirect(url_for("lots_list"))

    # Build sample set data for template
    sample_sets = lot.sample_sets.order_by(SampleSet.created_at).all()
    sample_set_data = []
    for ss in sample_sets:
        results = ss.test_results.order_by(TestResult.entered_at).all()
        sample_set_data.append({"set": ss, "results": results})

    audit_history = (
        AuditLog.query.filter_by(table_name="lots", record_id=lot_id)
        .union(
            AuditLog.query.filter_by(table_name="test_results").filter(
                AuditLog.details.ilike(f"%lot {lot_id}%")
            )
        )
        .union(
            AuditLog.query.filter_by(table_name="sample_sets").filter(
                AuditLog.details.ilike(f"%lot {lot_id}%")
            )
        )
        .order_by(AuditLog.timestamp.desc())
        .limit(50)
        .all()
    )

    reauth_users = User.query.filter_by(is_active_user=True, can_approve_edits=True).order_by(User.full_name).all()

    return render_template(
        "lot_detail.html",
        lot=lot,
        sample_set_data=sample_set_data,
        audit_history=audit_history,
        reauth_users=reauth_users,
    )


# ---------------------------------------------------------------------------
# Re-auth for editing completed lots
# ---------------------------------------------------------------------------


@app.route("/api/reauth-edit", methods=["POST"])
@login_required
def reauth_edit():
    csrf.protect()
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid request"}), 400

    username = data.get("username", "").strip()
    password = data.get("password", "")
    lot_id = data.get("lot_id")

    user = User.query.filter_by(username=username).first()
    if not user or not user.check_password(password) or not user.is_active_user:
        return jsonify({"error": "Invalid username or password."}), 401

    if not user.can_approve_edits:
        return jsonify({"error": "This user does not have approval permission."}), 403

    # Store authorization in session for this lot
    from flask import session
    session[f"reauth_lot_{lot_id}"] = user.id

    log_audit(
        "reauth",
        "lots",
        lot_id,
        f"Re-authenticated by {user.full_name} ({user.username}) to edit completed lot",
    )

    # Build redirect URL based on action
    action = data.get("action", "edit")
    sample_set_id = data.get("sample_set_id")
    if action == "add_set":
        redirect_url = url_for("add_sample_set", lot_id=lot_id)
    elif sample_set_id:
        redirect_url = url_for("test_entry", lot_id=lot_id, sample_set_id=sample_set_id)
    else:
        redirect_url = url_for("test_entry", lot_id=lot_id)

    return jsonify({"success": True, "redirect_url": redirect_url})


# ---------------------------------------------------------------------------
# Update lot status
# ---------------------------------------------------------------------------


@app.route("/lots/<int:lot_id>/status", methods=["POST"])
@login_required
def update_lot_status(lot_id):
    lot = db.session.get(Lot, lot_id)
    if not lot:
        flash("Lot not found.", "error")
        return redirect(url_for("lots_list"))

    old_status = lot.status
    new_status = request.form.get("status", "").strip()
    if new_status in ("pending", "in_progress", "completed", "flagged"):
        lot.status = new_status
        db.session.commit()
        log_audit(
            "update",
            "lots",
            lot.id,
            f"Status changed: {old_status} -> {new_status}",
        )
        flash(f"Status updated to '{new_status}'.", "success")

    return redirect(url_for("lot_detail", lot_id=lot.id))


# ---------------------------------------------------------------------------
# Add sample set
# ---------------------------------------------------------------------------


@app.route("/lots/<int:lot_id>/add-sample-set", methods=["POST"])
@login_required
def add_sample_set(lot_id):
    lot = db.session.get(Lot, lot_id)
    if not lot:
        flash("Lot not found.", "error")
        return redirect(url_for("lots_list"))

    label = lot.next_sample_set_label
    ss = SampleSet(
        lot_id=lot.id,
        label=label,
        created_by=current_user.id,
    )
    db.session.add(ss)

    # Set lot status to in_progress when adding a new set
    old_status = lot.status
    if lot.status == "completed":
        lot.status = "in_progress"

    db.session.commit()

    log_audit(
        "create",
        "sample_sets",
        ss.id,
        f"Created Sample Set {label} for lot {lot_id} (status: {old_status} -> {lot.status})",
    )

    flash(f"Sample Set {label} created!", "success")
    return redirect(url_for("test_entry", lot_id=lot.id, sample_set_id=ss.id))


# ---------------------------------------------------------------------------
# Test data entry
# ---------------------------------------------------------------------------


@app.route("/lots/<int:lot_id>/test-entry", methods=["GET", "POST"])
@login_required
def test_entry(lot_id):
    from flask import session

    lot = db.session.get(Lot, lot_id)
    if not lot:
        flash("Lot not found.", "error")
        return redirect(url_for("lots_list"))

    # Completed lots require re-authentication
    if lot.status == "completed":
        reauth_key = f"reauth_lot_{lot_id}"
        if reauth_key not in session:
            flash("Re-authentication required to edit a completed lot.", "error")
            return redirect(url_for("lot_detail", lot_id=lot.id))

    # Determine which sample set to work with
    sample_set_id = request.args.get("sample_set_id", type=int)

    if sample_set_id:
        sample_set = db.session.get(SampleSet, sample_set_id)
        if not sample_set or sample_set.lot_id != lot.id:
            flash("Sample set not found.", "error")
            return redirect(url_for("lot_detail", lot_id=lot.id))

        # Locked sample set check: non-latest sets require re-auth
        latest_set = SampleSet.query.filter_by(lot_id=lot.id).order_by(SampleSet.created_at.desc()).first()
        if latest_set and sample_set.id != latest_set.id:
            reauth_key = f"reauth_lot_{lot_id}"
            if reauth_key not in session:
                flash("Re-authentication required to edit a locked sample set.", "error")
                return redirect(url_for("lot_detail", lot_id=lot.id))
    else:
        set_count = lot.sample_set_count
        if set_count == 0:
            # Auto-create "Set A"
            sample_set = SampleSet(
                lot_id=lot.id,
                label="A",
                created_by=current_user.id,
            )
            db.session.add(sample_set)
            db.session.commit()
        elif set_count == 1:
            sample_set = lot.sample_sets.first()
        else:
            # Multiple sets — redirect to lot detail to pick one
            flash("Please select a sample set to edit.", "info")
            return redirect(url_for("lot_detail", lot_id=lot.id))

    # Get predefined parameters for this item
    parameters = lot.item.test_parameters.all()
    existing_results = {}
    existing_dynamic = []

    if parameters:
        # Map existing results by parameter name — scoped to this sample set
        for r in sample_set.test_results.all():
            existing_results[r.parameter_name] = r.value or ""
    else:
        existing_dynamic = sample_set.test_results.all()

    if request.method == "POST":
        # Save supervisor notification
        supervisor = request.form.get("notified_supervisor", "").strip()
        if supervisor:
            lot.notified_supervisor = supervisor

        # Clear old results for THIS SAMPLE SET only
        old_results = sample_set.test_results.all()
        old_data = [
            {"param": r.parameter_name, "value": r.value} for r in old_results
        ]
        for r in old_results:
            db.session.delete(r)

        new_results = []

        if parameters:
            # Predefined parameters
            count = int(request.form.get("param_count", 0))
            for i in range(1, count + 1):
                name = request.form.get(f"param_name_{i}", "").strip()
                value = request.form.get(f"param_value_{i}", "").strip()
                unit = request.form.get(f"param_unit_{i}", "").strip()
                min_v = request.form.get(f"param_min_{i}", "").strip()
                max_v = request.form.get(f"param_max_{i}", "").strip()

                if not name:
                    continue

                result = TestResult(
                    lot_id=lot.id,
                    sample_set_id=sample_set.id,
                    parameter_name=name,
                    value=value,
                    unit=unit or None,
                    min_spec=float(min_v) if min_v else None,
                    max_spec=float(max_v) if max_v else None,
                    entered_by=current_user.id,
                )
                # Try to set numeric value
                try:
                    result.numeric_value = float(value)
                except (ValueError, TypeError):
                    result.numeric_value = None

                result.evaluate_pass_fail()
                db.session.add(result)
                new_results.append({"param": name, "value": value})
        else:
            # Dynamic parameters
            dyn_count = int(request.form.get("dyn_count", 0))
            for i in range(1, dyn_count + 1):
                name = request.form.get(f"dyn_name_{i}", "").strip()
                value = request.form.get(f"dyn_value_{i}", "").strip()
                unit = request.form.get(f"dyn_unit_{i}", "").strip()
                min_v = request.form.get(f"dyn_min_{i}", "").strip()
                max_v = request.form.get(f"dyn_max_{i}", "").strip()

                if not name:
                    continue

                result = TestResult(
                    lot_id=lot.id,
                    sample_set_id=sample_set.id,
                    parameter_name=name,
                    value=value,
                    unit=unit or None,
                    min_spec=float(min_v) if min_v else None,
                    max_spec=float(max_v) if max_v else None,
                    entered_by=current_user.id,
                )
                try:
                    result.numeric_value = float(value)
                except (ValueError, TypeError):
                    result.numeric_value = None

                result.evaluate_pass_fail()
                db.session.add(result)
                new_results.append({"param": name, "value": value})

        db.session.commit()

        # Build audit detail with re-auth info if editing a completed lot
        reauth_key = f"reauth_lot_{lot_id}"
        reauth_user_id = session.get(reauth_key)
        if reauth_user_id:
            reauth_user = db.session.get(User, reauth_user_id)
            reauth_info = f" [COMPLETED LOT EDIT - authorized by {reauth_user.full_name} ({reauth_user.username})]"
            # Clear the session key so they must re-auth for future edits
            session.pop(reauth_key, None)
        else:
            reauth_info = ""

        detail = (
            f"Test results for lot {lot.id} Set {sample_set.label}{reauth_info}: "
            f"old={json.dumps(old_data)}, new={json.dumps(new_results)}"
        )
        log_audit("update", "test_results", lot.id, detail)

        # Auto-complete: if item has predefined parameters and every one
        # has a value filled in, set lot status to completed
        if parameters and lot.status != "completed":
            required_names = {p.parameter_name for p in parameters}
            filled_names = {r["param"] for r in new_results if r["value"]}
            if required_names <= filled_names:
                old_status = lot.status
                lot.status = "completed"
                db.session.commit()
                log_audit(
                    "update", "lots", lot.id,
                    f"Auto-completed: all spec fields filled ({old_status} -> completed)",
                )
                flash("Test results saved and lot auto-marked as completed!", "success")
            else:
                flash("Test results saved successfully!", "success")
        else:
            flash("Test results saved successfully!", "success")
        return redirect(url_for("lot_detail", lot_id=lot.id))

    supervisors = Supervisor.query.filter_by(is_active=True).order_by(Supervisor.name).all()

    return render_template(
        "test_entry.html",
        lot=lot,
        parameters=parameters,
        existing_results=existing_results,
        existing_dynamic=existing_dynamic,
        sample_set=sample_set,
        supervisors=supervisors,
    )


# ---------------------------------------------------------------------------
# Audit trail
# ---------------------------------------------------------------------------


@app.route("/audit")
@login_required
def audit_log():
    query = AuditLog.query
    users = User.query.order_by(User.full_name).all()

    # Filters
    user_id = request.args.get("user_id", "").strip()
    action = request.args.get("action", "").strip()
    table_name = request.args.get("table_name", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    if user_id:
        query = query.filter_by(user_id=int(user_id))
    if action:
        query = query.filter_by(action=action)
    if table_name:
        query = query.filter_by(table_name=table_name)
    if date_from:
        query = query.filter(
            AuditLog.timestamp >= datetime.strptime(date_from, "%Y-%m-%d")
        )
    if date_to:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59
        )
        query = query.filter(AuditLog.timestamp <= dt_to)

    # Build query string for pagination
    qs_parts = []
    for k in ("user_id", "action", "table_name", "date_from", "date_to"):
        v = request.args.get(k, "")
        if v:
            qs_parts.append(f"{k}={v}")
    query_string = "&".join(qs_parts)

    page = request.args.get("page", 1, type=int)
    pagination = query.order_by(AuditLog.timestamp.desc()).paginate(
        page=page, per_page=50, error_out=False
    )

    return render_template(
        "audit.html",
        logs=pagination.items,
        pagination=pagination,
        users=users,
        query_string=query_string,
    )


# ---------------------------------------------------------------------------
# Admin – Manage Items
# ---------------------------------------------------------------------------


@app.route("/admin/items", methods=["GET", "POST"])
@login_required
@lab_owner_or_admin_required
def manage_items():
    if request.method == "POST":
        action = request.form.get("action")

        if action in ("add", "edit"):
            item_number = request.form.get("item_number", "").strip()
            item_name = request.form.get("item_name", "").strip()
            description = request.form.get("description", "").strip()

            if action == "edit":
                item_id = int(request.form.get("item_id"))
                item = db.session.get(Item, item_id)
                if not item:
                    flash("Item not found.", "error")
                    return redirect(url_for("manage_items"))
                item.item_number = item_number
                item.item_name = item_name
                item.description = description

                # Clear old params and re-add
                TestParameter.query.filter_by(item_id=item.id).delete()
                audit_action = "update"
            else:
                # Check for duplicate item number
                if Item.query.filter_by(item_number=item_number).first():
                    flash(f"Item number '{item_number}' already exists.", "error")
                    return redirect(url_for("manage_items"))

                item = Item(
                    item_number=item_number,
                    item_name=item_name,
                    description=description,
                )
                db.session.add(item)
                db.session.flush()  # Get item.id
                audit_action = "create"

            # Add test parameters
            tp_names = request.form.getlist("tp_name")
            tp_units = request.form.getlist("tp_unit")
            tp_mins = request.form.getlist("tp_min")
            tp_maxs = request.form.getlist("tp_max")

            for i in range(len(tp_names)):
                name = tp_names[i].strip()
                if not name:
                    continue
                tp = TestParameter(
                    item_id=item.id,
                    parameter_name=name,
                    unit=tp_units[i].strip() if i < len(tp_units) else None,
                    min_value=(
                        float(tp_mins[i]) if i < len(tp_mins) and tp_mins[i].strip() else None
                    ),
                    max_value=(
                        float(tp_maxs[i]) if i < len(tp_maxs) and tp_maxs[i].strip() else None
                    ),
                )
                db.session.add(tp)

            db.session.commit()
            log_audit(audit_action, "items", item.id, f"Item '{item_number}' {audit_action}d")
            flash(f"Item '{item_number}' {audit_action}d successfully!", "success")
            return redirect(url_for("manage_items"))

    items = Item.query.order_by(Item.item_number).all()
    templates = TestTemplate.query.order_by(TestTemplate.name).all()

    edit_item = None
    edit_id = request.args.get("edit")
    if edit_id:
        edit_item = db.session.get(Item, int(edit_id))

    return render_template("manage_items.html", items=items, edit_item=edit_item, templates=templates)


# ---------------------------------------------------------------------------
# Admin – Manage Users
# ---------------------------------------------------------------------------


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@lab_owner_or_admin_required
def manage_users():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "add":
            username = request.form.get("username", "").strip()
            full_name = request.form.get("full_name", "").strip()
            password = request.form.get("password", "")
            role = request.form.get("role", "operator")
            can_approve = request.form.get("can_approve_edits") == "on"

            # Lab owners cannot create admin users
            if not current_user.is_admin and role == "admin":
                flash("Only admins can create admin users.", "error")
                return redirect(url_for("manage_users"))

            if User.query.filter_by(username=username).first():
                flash(f"Username '{username}' already exists.", "error")
                return redirect(url_for("manage_users"))

            user = User(username=username, full_name=full_name, role=role, can_approve_edits=can_approve)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            log_audit("create", "users", user.id, f"User '{username}' created (role: {role}, can_approve_edits: {can_approve})")
            flash(f"User '{username}' created successfully!", "success")

        elif action == "toggle":
            user_id = int(request.form.get("user_id"))
            user = db.session.get(User, user_id)
            if user and user.id != current_user.id:
                user.is_active_user = not user.is_active_user
                db.session.commit()
                status = "activated" if user.is_active_user else "deactivated"
                log_audit("update", "users", user.id, f"User '{user.username}' {status}")
                flash(f"User '{user.username}' {status}.", "success")

        elif action == "reset_password":
            user_id = int(request.form.get("user_id"))
            new_password = request.form.get("new_password", "password")
            user = db.session.get(User, user_id)
            if user:
                user.set_password(new_password)
                db.session.commit()
                log_audit("update", "users", user.id, f"Password reset for '{user.username}'")
                flash(f"Password reset for '{user.username}'.", "success")

        elif action == "toggle_approve":
            user_id = int(request.form.get("user_id"))
            user = db.session.get(User, user_id)
            if user and user.id != current_user.id:
                user.can_approve_edits = not user.can_approve_edits
                db.session.commit()
                state = "granted" if user.can_approve_edits else "revoked"
                log_audit("update", "users", user.id, f"Approval permission {state} for '{user.username}'")
                flash(f"Approval permission {state} for '{user.username}'.", "success")

        return redirect(url_for("manage_users"))

    users = User.query.order_by(User.full_name).all()
    return render_template("manage_users.html", users=users)


# ---------------------------------------------------------------------------
# Lab Manager – Supervisors
# ---------------------------------------------------------------------------


@app.route("/admin/supervisors", methods=["GET", "POST"])
@login_required
@lab_owner_or_admin_required
def manage_supervisors():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "add":
            name = request.form.get("name", "").strip()
            if not name:
                flash("Supervisor name is required.", "error")
            elif Supervisor.query.filter_by(name=name).first():
                flash(f"Supervisor '{name}' already exists.", "error")
            else:
                sup = Supervisor(name=name)
                db.session.add(sup)
                db.session.commit()
                log_audit("create", "supervisors", sup.id, f"Supervisor '{name}' added")
                flash(f"Supervisor '{name}' added successfully!", "success")

        elif action == "edit":
            sup_id = int(request.form.get("supervisor_id"))
            new_name = request.form.get("name", "").strip()
            sup = db.session.get(Supervisor, sup_id)
            if sup and new_name:
                existing = Supervisor.query.filter(Supervisor.name == new_name, Supervisor.id != sup_id).first()
                if existing:
                    flash(f"Supervisor '{new_name}' already exists.", "error")
                else:
                    old_name = sup.name
                    sup.name = new_name
                    db.session.commit()
                    log_audit("update", "supervisors", sup.id, f"Supervisor renamed from '{old_name}' to '{new_name}'")
                    flash(f"Supervisor renamed to '{new_name}'.", "success")

        elif action == "toggle":
            sup_id = int(request.form.get("supervisor_id"))
            sup = db.session.get(Supervisor, sup_id)
            if sup:
                sup.is_active = not sup.is_active
                db.session.commit()
                status = "activated" if sup.is_active else "deactivated"
                log_audit("update", "supervisors", sup.id, f"Supervisor '{sup.name}' {status}")
                flash(f"Supervisor '{sup.name}' {status}.", "success")

        elif action == "delete":
            sup_id = int(request.form.get("supervisor_id"))
            sup = db.session.get(Supervisor, sup_id)
            if sup:
                name = sup.name
                db.session.delete(sup)
                db.session.commit()
                log_audit("delete", "supervisors", sup_id, f"Supervisor '{name}' deleted")
                flash(f"Supervisor '{name}' deleted.", "success")

        return redirect(url_for("manage_supervisors"))

    supervisors = Supervisor.query.order_by(Supervisor.name).all()
    return render_template("manage_supervisors.html", supervisors=supervisors)


# ---------------------------------------------------------------------------
# Lab Manager – Test Templates
# ---------------------------------------------------------------------------


@app.route("/admin/templates", methods=["GET", "POST"])
@login_required
@lab_owner_or_admin_required
def manage_templates():
    if request.method == "POST":
        action = request.form.get("action")

        if action in ("add", "edit"):
            name = request.form.get("name", "").strip()
            description = request.form.get("description", "").strip()

            if action == "edit":
                template_id = int(request.form.get("template_id"))
                template = db.session.get(TestTemplate, template_id)
                if not template:
                    flash("Template not found.", "error")
                    return redirect(url_for("manage_templates"))
                template.name = name
                template.description = description

                # Clear old params and re-add
                TemplateParameter.query.filter_by(template_id=template.id).delete()
                audit_action = "update"
            else:
                if TestTemplate.query.filter_by(name=name).first():
                    flash(f"Template '{name}' already exists.", "error")
                    return redirect(url_for("manage_templates"))

                template = TestTemplate(
                    name=name,
                    description=description,
                    created_by=current_user.id,
                )
                db.session.add(template)
                db.session.flush()
                audit_action = "create"

            # Add parameters
            tp_names = request.form.getlist("tp_name")
            tp_units = request.form.getlist("tp_unit")
            tp_mins = request.form.getlist("tp_min")
            tp_maxs = request.form.getlist("tp_max")

            for i in range(len(tp_names)):
                pname = tp_names[i].strip()
                if not pname:
                    continue
                tp = TemplateParameter(
                    template_id=template.id,
                    parameter_name=pname,
                    unit=tp_units[i].strip() if i < len(tp_units) else None,
                    min_value=(
                        float(tp_mins[i]) if i < len(tp_mins) and tp_mins[i].strip() else None
                    ),
                    max_value=(
                        float(tp_maxs[i]) if i < len(tp_maxs) and tp_maxs[i].strip() else None
                    ),
                )
                db.session.add(tp)

            db.session.commit()
            log_audit(audit_action, "test_templates", template.id, f"Template '{name}' {audit_action}d")
            flash(f"Template '{name}' {audit_action}d successfully!", "success")
            return redirect(url_for("manage_templates"))

        elif action == "delete":
            template_id = int(request.form.get("template_id"))
            template = db.session.get(TestTemplate, template_id)
            if template:
                tname = template.name
                db.session.delete(template)
                db.session.commit()
                log_audit("delete", "test_templates", template_id, f"Template '{tname}' deleted")
                flash(f"Template '{tname}' deleted.", "success")
            return redirect(url_for("manage_templates"))

    templates = TestTemplate.query.order_by(TestTemplate.name).all()

    edit_template = None
    edit_id = request.args.get("edit")
    if edit_id:
        edit_template = db.session.get(TestTemplate, int(edit_id))

    return render_template("manage_templates.html", templates=templates, edit_template=edit_template)


@app.route("/api/template/<int:template_id>")
@login_required
@lab_owner_or_admin_required
def get_template_params(template_id):
    template = db.session.get(TestTemplate, template_id)
    if not template:
        return jsonify({"error": "Template not found"}), 404

    params = [
        {
            "name": p.parameter_name,
            "unit": p.unit or "",
            "min": p.min_value,
            "max": p.max_value,
        }
        for p in template.parameters.all()
    ]
    return jsonify({"name": template.name, "parameters": params})


# ---------------------------------------------------------------------------
# Feedback / Ticket System
# ---------------------------------------------------------------------------


@app.route("/api/submit-ticket", methods=["POST"])
@login_required
def submit_ticket():
    csrf.protect()
    data = request.get_json()
    if not data or not data.get("description", "").strip():
        return jsonify({"error": "Description is required"}), 400

    ticket = SystemRequest(
        user_id=current_user.id,
        description=data["description"].strip(),
        screenshot_data=data.get("screenshot"),
        page_url=data.get("page_url", ""),
    )
    db.session.add(ticket)
    db.session.commit()
    log_audit("create", "system_requests", ticket.id, "Feedback ticket submitted")
    return jsonify({"success": True, "ticket_id": ticket.id})


@app.route("/admin/tickets")
@login_required
@admin_required
def manage_tickets():
    query = SystemRequest.query

    status_filter = request.args.get("status", "").strip()
    if status_filter:
        query = query.filter_by(status=status_filter)

    page = request.args.get("page", 1, type=int)
    pagination = query.order_by(SystemRequest.created_at.desc()).paginate(
        page=page, per_page=25, error_out=False
    )

    # Counts for filter badges
    counts = {
        "all": SystemRequest.query.count(),
        "new": SystemRequest.query.filter_by(status="new").count(),
        "reviewed": SystemRequest.query.filter_by(status="reviewed").count(),
        "in_progress": SystemRequest.query.filter_by(status="in_progress").count(),
        "completed": SystemRequest.query.filter_by(status="completed").count(),
        "dismissed": SystemRequest.query.filter_by(status="dismissed").count(),
    }

    return render_template(
        "manage_tickets.html",
        tickets=pagination.items,
        pagination=pagination,
        counts=counts,
        current_status=status_filter,
    )


@app.route("/admin/tickets/<int:ticket_id>", methods=["GET", "POST"])
@login_required
@admin_required
def ticket_detail(ticket_id):
    ticket = db.session.get(SystemRequest, ticket_id)
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("manage_tickets"))

    if request.method == "POST":
        new_status = request.form.get("status", "").strip()
        admin_notes = request.form.get("admin_notes", "").strip()

        if new_status in ("new", "reviewed", "in_progress", "completed", "dismissed"):
            old_status = ticket.status
            ticket.status = new_status
            ticket.admin_notes = admin_notes
            db.session.commit()
            log_audit(
                "update",
                "system_requests",
                ticket.id,
                f"Ticket status: {old_status} -> {new_status}",
            )
            flash("Ticket updated.", "success")
        return redirect(url_for("ticket_detail", ticket_id=ticket.id))

    return render_template("ticket_detail.html", ticket=ticket)


# ---------------------------------------------------------------------------
# Initialize DB & run
# ---------------------------------------------------------------------------


def init_db():
    """Create tables and seed an admin user if the DB is empty."""
    with app.app_context():
        db.create_all()

        # --- Migration: add sample_set_id column if missing ---
        from sqlalchemy import inspect as sa_inspect, text

        inspector = sa_inspect(db.engine)
        if "test_results" in inspector.get_table_names():
            columns = [c["name"] for c in inspector.get_columns("test_results")]
            if "sample_set_id" not in columns:
                db.session.execute(
                    text("ALTER TABLE test_results ADD COLUMN sample_set_id INTEGER REFERENCES sample_sets(id)")
                )
                db.session.commit()

        # --- Migration: add can_approve_edits column if missing ---
        if "users" in inspector.get_table_names():
            user_columns = [c["name"] for c in inspector.get_columns("users")]
            if "can_approve_edits" not in user_columns:
                db.session.execute(
                    text("ALTER TABLE users ADD COLUMN can_approve_edits BOOLEAN DEFAULT 0")
                )
                # Set True for admin and lab_owner users
                db.session.execute(
                    text("UPDATE users SET can_approve_edits = 1 WHERE role IN ('admin', 'lab_owner')")
                )
                db.session.commit()

        # --- Migration: create Set A for orphaned test results ---
        orphaned = (
            TestResult.query.filter(TestResult.sample_set_id.is_(None)).all()
        )
        if orphaned:
            # Group by lot_id
            lots_with_orphans = {}
            for tr in orphaned:
                lots_with_orphans.setdefault(tr.lot_id, []).append(tr)

            for lot_id, results in lots_with_orphans.items():
                lot = db.session.get(Lot, lot_id)
                if not lot:
                    continue
                # Only create if no sample sets exist yet
                if lot.sample_sets.count() == 0:
                    ss = SampleSet(
                        lot_id=lot_id,
                        label="A",
                        created_by=lot.created_by,
                    )
                    db.session.add(ss)
                    db.session.flush()
                else:
                    ss = lot.sample_sets.first()

                for tr in results:
                    tr.sample_set_id = ss.id

            db.session.commit()
            print(f"Migration: linked {len(orphaned)} orphaned test results to sample sets.")

        # --- Migration: seed default supervisors if table is empty ---
        if not Supervisor.query.first():
            default_supervisors = [
                "Adam Sella",
                "Garry Hargis",
                "Jeff Freeman",
                "Wayne Nadeau",
            ]
            for name in default_supervisors:
                db.session.add(Supervisor(name=name))
            db.session.commit()
            print("Migration: seeded default supervisors.")

        if not User.query.first():
            admin = User(
                username="admin",
                full_name="System Administrator",
                role="admin",
                can_approve_edits=True,
            )
            admin.set_password("password")
            db.session.add(admin)

            # Seed a sample operator
            operator = User(
                username="operator1",
                full_name="Lab Operator",
                role="operator",
            )
            operator.set_password("password")
            db.session.add(operator)

            # Seed a lab manager
            lab_mgr = User(
                username="labowner",
                full_name="Lab Owner",
                role="lab_owner",
                can_approve_edits=True,
            )
            lab_mgr.set_password("password")
            db.session.add(lab_mgr)

            # Seed a viewer
            viewer = User(
                username="viewer1",
                full_name="QC Viewer",
                role="viewer",
            )
            viewer.set_password("password")
            db.session.add(viewer)

            # Seed sample items
            item1 = Item(
                item_number="ITEM-001",
                item_name="Raw Material A",
                description="Primary raw material for production",
            )
            db.session.add(item1)
            db.session.flush()

            tp1 = TestParameter(
                item_id=item1.id,
                parameter_name="pH",
                unit="",
                min_value=6.0,
                max_value=8.0,
            )
            tp2 = TestParameter(
                item_id=item1.id,
                parameter_name="Viscosity",
                unit="cP",
                min_value=100,
                max_value=500,
            )
            tp3 = TestParameter(
                item_id=item1.id,
                parameter_name="Moisture Content",
                unit="%",
                min_value=0,
                max_value=5.0,
            )
            db.session.add_all([tp1, tp2, tp3])

            item2 = Item(
                item_number="ITEM-002",
                item_name="Finished Product B",
                description="Final packaged product",
            )
            db.session.add(item2)
            db.session.flush()

            tp4 = TestParameter(
                item_id=item2.id,
                parameter_name="Weight",
                unit="g",
                min_value=95.0,
                max_value=105.0,
            )
            tp5 = TestParameter(
                item_id=item2.id,
                parameter_name="Color Score",
                unit="L*",
                min_value=80,
                max_value=100,
            )
            db.session.add_all([tp4, tp5])
            db.session.flush()

            # --- Seed fake lots with test results ---
            from random import uniform, choice, randint

            lot_data = [
                # (lot_number, item, status, is_dup)
                ("LOT-2024-001", item1, "completed", False),
                ("LOT-2024-002", item1, "completed", False),
                ("LOT-2024-003", item2, "completed", False),
                ("LOT-2024-004", item1, "in_progress", False),
                ("LOT-2024-005", item2, "completed", False),
                ("LOT-2024-006", item1, "flagged", False),
                ("LOT-2024-007", item2, "pending", False),
                ("LOT-2024-008", item1, "completed", False),
                ("LOT-2024-009", item2, "in_progress", False),
                ("LOT-2024-010", item1, "completed", False),
                ("LOT-2024-011", item2, "completed", False),
                ("LOT-2024-012", item1, "pending", False),
                ("LOT-2024-013", item2, "flagged", False),
                ("LOT-2024-014", item1, "completed", False),
                ("LOT-2024-015", item2, "completed", False),
                ("LOT-2024-001", item1, "in_progress", True),  # duplicate
                ("LOT-2024-016", item1, "completed", False),
                ("LOT-2024-017", item2, "pending", False),
                ("LOT-2024-018", item1, "completed", False),
                ("LOT-2024-019", item2, "completed", False),
            ]

            created_lots = []
            users_for_lots = [admin, operator]
            base_date = date(2024, 6, 1)

            for i, (lot_num, item, status, is_dup) in enumerate(lot_data):
                recv_date = date(2024, 6 + i // 5, 1 + (i * 3) % 28)
                creator = choice(users_for_lots)
                lot = Lot(
                    lot_number=lot_num,
                    barcode=lot_num,
                    item_id=item.id,
                    date_received=recv_date,
                    status=status,
                    is_duplicate=is_dup,
                    duplicate_of_id=created_lots[0].id if is_dup and created_lots else None,
                    notes=f"Sample lot for testing" if not is_dup else "Duplicate lot for testing",
                    created_by=creator.id,
                )
                db.session.add(lot)
                db.session.flush()
                created_lots.append(lot)

                # Add test results for completed / in_progress / flagged lots
                if status in ("completed", "in_progress", "flagged"):
                    # Create a sample set for these results
                    sample_set = SampleSet(
                        lot_id=lot.id,
                        label="A",
                        created_by=creator.id,
                    )
                    db.session.add(sample_set)
                    db.session.flush()

                    params = item.test_parameters.all()
                    for param in params:
                        # Generate a value — mostly in range, sometimes out for flagged
                        if status == "flagged" and param == params[0]:
                            # Force a failing value for the first param
                            if param.min_value is not None and param.max_value is not None:
                                val = param.max_value + uniform(1, 5)
                            else:
                                val = 999.0
                        else:
                            lo = param.min_value if param.min_value is not None else 0
                            hi = param.max_value if param.max_value is not None else 100
                            val = round(uniform(lo, hi), 2)

                        result = TestResult(
                            lot_id=lot.id,
                            sample_set_id=sample_set.id,
                            parameter_name=param.parameter_name,
                            value=str(val),
                            numeric_value=val,
                            unit=param.unit,
                            min_spec=param.min_value,
                            max_spec=param.max_value,
                            entered_by=creator.id,
                        )
                        result.evaluate_pass_fail()
                        db.session.add(result)

            db.session.commit()
            print("Database initialized with default admin user, sample items, and fake lots.")
        else:
            print("Database already initialized.")


# ---------------------------------------------------------------------------
# Integration API – cross-system endpoints (API-key auth, no login required)
# ---------------------------------------------------------------------------


def require_integration_key(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        key = request.headers.get("X-Integration-Key")
        expected = app.config.get(
            "INTEGRATION_API_KEY", os.environ.get("INTEGRATION_API_KEY", "")
        )
        if not key or key != expected:
            return jsonify({"success": False, "error": "Invalid or missing integration key"}), 401
        return f(*args, **kwargs)
    return decorated


@app.route("/api/integration/health")
@csrf.exempt
@require_integration_key
def integration_health():
    return jsonify({"success": True, "app": "qc-lab", "version": "1.0", "status": "ok"})


@app.route("/api/integration/lots")
@csrf.exempt
@require_integration_key
def integration_lots():
    search = request.args.get("search", "")
    status = request.args.get("status", "")
    limit = min(int(request.args.get("limit", 20)), 100)

    query = Lot.query.join(Item)
    if search:
        query = query.filter(Lot.lot_number.ilike(f"%{search}%"))
    if status:
        query = query.filter(Lot.status == status)

    lots = query.order_by(Lot.created_at.desc()).limit(limit).all()

    return jsonify({
        "success": True,
        "data": [{
            "id": lot.id,
            "lot_number": lot.lot_number,
            "item_name": lot.item.item_name if lot.item else None,
            "item_number": lot.item.item_number if lot.item else None,
            "status": lot.status,
            "is_passing": lot.is_passing,
            "machine": lot.machine,
            "created_at": lot.created_at.isoformat() if lot.created_at else None,
        } for lot in lots],
    })


@app.route("/api/integration/lots/<int:lot_id>")
@csrf.exempt
@require_integration_key
def integration_lot_detail(lot_id):
    lot = Lot.query.get_or_404(lot_id)
    results = TestResult.query.filter_by(lot_id=lot.id).all()
    passing = sum(1 for r in results if r.is_passing)
    failing = sum(1 for r in results if r.is_passing is False)

    return jsonify({
        "success": True,
        "data": {
            "id": lot.id,
            "lot_number": lot.lot_number,
            "item_name": lot.item.item_name if lot.item else None,
            "item_number": lot.item.item_number if lot.item else None,
            "status": lot.status,
            "is_passing": lot.is_passing,
            "machine": lot.machine,
            "sample_sets": lot.sample_set_count,
            "test_results_summary": {
                "total": len(results),
                "passing": passing,
                "failing": failing,
            },
            "created_at": lot.created_at.isoformat() if lot.created_at else None,
            "notes": lot.notes,
        },
    })


@app.after_request
def add_integration_cors(response):
    if request.path.startswith("/api/integration/"):
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Integration-Key"
        response.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    return response


if __name__ == "__main__":
    init_db()

    debug = os.environ.get("FLASK_ENV") != "production"
    app.run(debug=debug, host="0.0.0.0", port=5000)

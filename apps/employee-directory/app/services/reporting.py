from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..config import Config
from ..models import Company, Employee, LifecycleEvent


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (
            length + 2
        )


def export_asset_snapshot(export_dir: Path | None = None) -> Path:
    export_dir = export_dir or Config.EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    export_path = export_dir / f"employee_assets_{timestamp}.xlsx"

    wb = Workbook()

    # Employees tab
    ws_employees = wb.active
    ws_employees.title = "Employees"
    ws_employees.append(
        [
            "Employee ID",
            "Employee Number",
            "First Name",
            "Last Name",
            "Preferred Name",
            "Email",
            "Department",
            "Title",
            "Status",
            "Start Date",
            "Termination Date",
        ]
    )
    for employee in Employee.query.order_by(Employee.last_name).all():
        ws_employees.append(
            [
                employee.id,
                employee.employee_number,
                employee.first_name,
                employee.last_name,
                employee.preferred_name,
                employee.email,
                employee.department,
                employee.title,
                employee.status.value,
                employee.start_date.isoformat() if employee.start_date else None,
                employee.termination_date.isoformat()
                if employee.termination_date
                else None,
            ]
        )
    _autosize_columns(ws_employees)

    # Hardware tab
    ws_hardware = wb.create_sheet("Hardware")
    ws_hardware.append(
        [
            "Asset ID",
            "Type",
            "Manufacturer",
            "Model",
            "Serial Number",
            "Asset Tag",
            "Status",
            "Assigned Employee ID",
            "Assigned Date",
            "Return Due Date",
        ]
    )
    for employee in Employee.query.all():
        for asset in employee.hardware_assets:
            ws_hardware.append(
                [
                    asset.id,
                    asset.asset_type,
                    asset.manufacturer,
                    asset.model,
                    asset.serial_number,
                    asset.asset_tag,
                    asset.status.value,
                    employee.id,
                    asset.assigned_date.isoformat()
                    if asset.assigned_date
                    else None,
                    asset.return_due_date.isoformat()
                    if asset.return_due_date
                    else None,
                ]
            )
    _autosize_columns(ws_hardware)

    # Software tab
    ws_software = wb.create_sheet("Software")
    ws_software.append(
        [
            "Subscription ID",
            "Name",
            "Vendor",
            "License Identifier",
            "Billing Cycle",
            "Cost",
            "Status",
            "Assigned Employee ID",
            "Renewal Date",
        ]
    )
    for employee in Employee.query.all():
        for sub in employee.software_subscriptions:
            ws_software.append(
                [
                    sub.id,
                    sub.name,
                    sub.vendor,
                    sub.license_identifier,
                    sub.billing_cycle,
                    sub.cost,
                    sub.status.value,
                    employee.id,
                    sub.renewal_date.isoformat() if sub.renewal_date else None,
                ]
            )
    _autosize_columns(ws_software)

    # Directory Groups tab
    ws_groups = wb.create_sheet("DirectoryGroups")
    ws_groups.append(
        [
            "Employee ID",
            "Employee Name",
            "Group Name",
            "Scope",
            "Type",
            "Source",
            "Description",
        ]
    )
    for employee in Employee.query.order_by(Employee.last_name).all():
        for group in employee.directory_groups:
            ws_groups.append(
                [
                    employee.id,
                    employee.full_name,
                    group.group_name,
                    group.group_scope,
                    group.group_type,
                    group.source.value,
                    group.description,
                ]
            )
    _autosize_columns(ws_groups)

    # Lifecycle tab
    ws_lifecycle = wb.create_sheet("LifecycleTasks")
    ws_lifecycle.append(
        [
            "Event ID",
            "Employee ID",
            "Event Type",
            "Event Status",
            "Task ID",
            "Task Type",
            "Task Description",
            "Task Status",
            "Due Date",
            "Completed At",
        ]
    )
    events = (
        LifecycleEvent.query.order_by(LifecycleEvent.created_at)
        .join(Employee)
        .all()
    )
    for event in events:
        for task in event.tasks:
            ws_lifecycle.append(
                [
                    event.id,
                    event.employee_id,
                    event.event_type.value,
                    event.status.value,
                    task.id,
                    task.task_type,
                    task.description,
                    task.status.value,
                    task.due_date.isoformat() if task.due_date else None,
                    task.completed_at.isoformat() if task.completed_at else None,
                ]
            )
    _autosize_columns(ws_lifecycle)

    wb.save(export_path)
    return export_path


_HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_SECTION_FILL = PatternFill(start_color="FFF5C542", end_color="FFF5C542", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="FF1F2937")


def _style_header_row(worksheet, row_number: int = 1) -> None:
    for cell in worksheet[row_number]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[row_number].height = 22
    worksheet.freeze_panes = worksheet.cell(row=row_number + 1, column=1).coordinate


def _visible_employees() -> list[Employee]:
    return (
        Employee.query.filter(Employee.primary_employee_id.is_(None))
        .order_by(Employee.last_name, Employee.first_name)
        .all()
    )


def _company_name_map() -> dict[int, str]:
    return {c.id: c.name for c in Company.query.all()}


def build_user_list_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = [
        "Last name",
        "First name",
        "Preferred name",
        "Email",
        "Title",
        "Department",
        "Company",
        "Status",
        "Account type",
        "Manager",
        "Start date",
        "Office phone",
        "Cell phone",
        "Extension",
        "Office location",
        "Employee #",
        "Alternate emails",
    ]
    ws.append(headers)
    _style_header_row(ws)

    companies = _company_name_map()
    for emp in _visible_employees():
        alts = ", ".join(emp.alternate_emails or [])
        ws.append(
            [
                emp.last_name,
                emp.first_name,
                emp.preferred_name,
                emp.email,
                emp.title,
                emp.department,
                companies.get(emp.company_id) if emp.company_id else None,
                emp.status.value,
                emp.account_type or "domain",
                emp.manager_email,
                emp.start_date.isoformat() if emp.start_date else None,
                emp.phone,
                emp.mobile_phone,
                emp.extension,
                emp.office_location,
                emp.employee_number,
                alts or None,
            ]
        )

    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_ROOMS_INFRA: list[tuple[str, str, str, str | None]] = [
    ("395", "Large Conference Room", "Office", None),
    ("352", "Production Supervisors", "Plant", "603-315-6110"),
    ("340", "Power House", "Plant", None),
    ("346", "QC Lab", "Plant", None),
    ("424", "Waste Area", "Plant", None),
    ("302", "Security", "Security office, front building", "603-493-4928"),
    ("313", "Server Room", "Office", None),
    ("343", "Shipping Lead (Les Magoon)", "Shipping", None),
    ("320", "Small Conference Room", "Office speaker phone", None),
]

_OUTSIDE_SUPPORT: list[tuple[str, str, str]] = [
    ("White Mountain", "Computer Support", "603-889-2210 / support@whitemtn.com"),
    ("PHD Communications", "Phone Line Support", "603-666-5533 / support@phdcom.com"),
]

_FOOTER_LINES = [
    "Main Phone #: 603-627-5150  |  Main Fax #: 603-627-5154",
    "Customer Service Phone #: 1-800-851-2001  |  Customer Service Fax #: 603-627-4499",
    "HR Fax #: 603-644-7658",
]


def _is_mexico_rep(e: Employee) -> bool:
    cell = (e.mobile_phone or "").strip()
    return cell.startswith("+") and "52" in cell.replace(" ", "")[:4]


def build_phone_directory_pdf() -> bytes:
    """PDF phone directory that mirrors the Acme Industries extension list layout."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.4 * inch,
        rightMargin=0.4 * inch,
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
        title="Acme Industries Phone Extension List",
    )

    sheet = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "dir_cell",
        parent=sheet["BodyText"],
        fontSize=8,
        leading=10,
        alignment=1,
    )
    footer_style = ParagraphStyle(
        "dir_footer",
        parent=sheet["BodyText"],
        fontSize=8,
        leading=11,
        alignment=1,
    )

    def p(text: str | None):
        return Paragraph(text or "", cell_style)

    eligible = [
        e
        for e in _visible_employees()
        if e.status.value == "active"
        and (e.account_type or "domain") == "domain"
        and (
            (e.phone or "").strip()
            or (e.mobile_phone or "").strip()
            or (e.extension or "").strip()
        )
    ]
    eligible.sort(key=lambda e: ((e.last_name or "").lower(), (e.first_name or "").lower()))

    mexico = [e for e in eligible if _is_mexico_rep(e)]
    main = [e for e in eligible if not _is_mexico_rep(e)]

    data: list[list] = []
    cmds: list[tuple] = []

    data.append(["Acme Industries Phone Extension List", "", "", ""])
    cmds += [
        ("SPAN", (0, 0), (3, 0)),
        ("BACKGROUND", (0, 0), (3, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (3, 0), colors.white),
        ("FONTNAME", (0, 0), (3, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (3, 0), 14),
        ("ALIGN", (0, 0), (3, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (3, 0), 6),
        ("BOTTOMPADDING", (0, 0), (3, 0), 6),
    ]

    data.append(["Ext", "Name", "Title", "Cell Phone"])
    cmds += [
        ("BACKGROUND", (0, 1), (3, 1), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 1), (3, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 1), (0, 1), "CENTER"),
    ]

    def _cell_display(e: Employee) -> str:
        parts = []
        if (e.mobile_phone or "").strip():
            parts.append(f"C: {e.mobile_phone.strip()}")
        if (e.phone or "").strip():
            parts.append(f"O: {e.phone.strip()}")
        return " / ".join(parts)

    for e in main:
        ext = e.extension or "offsite"
        name = f"{e.last_name}, {e.first_name}"
        data.append([ext, p(name), p(e.title), p(_cell_display(e))])

    def _section_header(label: str) -> None:
        data.append([label, "", "", ""])
        row = len(data) - 1
        cmds.extend(
            [
                ("SPAN", (0, row), (3, row)),
                ("BACKGROUND", (0, row), (3, row), colors.HexColor("#F5C542")),
                ("FONTNAME", (0, row), (3, row), "Helvetica-Bold"),
                ("ALIGN", (0, row), (3, row), "CENTER"),
                ("TOPPADDING", (0, row), (3, row), 4),
                ("BOTTOMPADDING", (0, row), (3, row), 4),
            ]
        )

    if mexico:
        _section_header("Mexico")
        for e in mexico:
            name = f"{e.last_name}, {e.first_name}"
            data.append(["Mexico", p(name), p(e.title), p(_cell_display(e))])

    _section_header("Rooms & Infrastructure")
    for ext, name, desc, phone in _ROOMS_INFRA:
        data.append([ext, p(name), p(desc), p(phone)])

    _section_header("Outside Support")
    for name, desc, contact in _OUTSIDE_SUPPORT:
        data.append(["", p(name), p(desc), p(contact)])

    cmds.extend(
        [
            ("GRID", (0, 1), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 2), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 2), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 2), (-1, -1), 2),
        ]
    )

    col_widths = [0.9 * inch, 2.2 * inch, 3.0 * inch, 1.6 * inch]
    tbl = Table(data, colWidths=col_widths, repeatRows=2)
    tbl.setStyle(TableStyle(cmds))

    story = [tbl, Spacer(1, 10)]
    for line in _FOOTER_LINES:
        story.append(Paragraph(line, footer_style))
    story.append(Spacer(1, 4))
    story.append(
        Paragraph(
            f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            footer_style,
        )
    )

    doc.build(story)
    return buf.getvalue()


def build_phone_directory_xlsx() -> bytes:
    """Directory of employees who have at least one of: office phone, cell, extension.

    Grouped by company → department, sorted by last name within each group.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Directory"

    headers = [
        "Name",
        "Title",
        "Department",
        "Extension",
        "Office phone",
        "Cell phone",
        "Email",
        "Office location",
    ]
    ws.append(headers)
    _style_header_row(ws)

    companies = _company_name_map()

    def _has_contact(e: Employee) -> bool:
        return bool((e.phone or "").strip() or (e.mobile_phone or "").strip() or (e.extension or "").strip())

    eligible = [
        e
        for e in _visible_employees()
        if _has_contact(e)
        and e.status.value == "active"
        and (e.account_type or "domain") == "domain"
    ]

    def _sort_key(e: Employee) -> tuple[str, str, str, str]:
        return (
            companies.get(e.company_id, "") if e.company_id else "zz",
            (e.department or "").lower(),
            (e.last_name or "").lower(),
            (e.first_name or "").lower(),
        )

    eligible.sort(key=_sort_key)

    current_company: str | None = object()  # type: ignore[assignment]
    current_dept: str | None = object()  # type: ignore[assignment]
    section_rows: list[int] = []

    for emp in eligible:
        company_name = companies.get(emp.company_id) if emp.company_id else "Unassigned"
        dept = emp.department or "(No department)"
        if company_name != current_company:
            ws.append([f"{company_name}"])
            section_rows.append(ws.max_row)
            current_company = company_name
            current_dept = object()  # force dept heading re-print
        if dept != current_dept:
            ws.append([f"  {dept}"])
            section_rows.append(ws.max_row)
            current_dept = dept

        name = emp.preferred_name or emp.full_name
        ws.append(
            [
                name,
                emp.title,
                emp.department,
                emp.extension,
                emp.phone,
                emp.mobile_phone,
                emp.email,
                emp.office_location,
            ]
        )

    # Style section heading rows
    for row_idx in section_rows:
        cell = ws.cell(row=row_idx, column=1)
        cell.fill = _SECTION_FILL
        cell.font = _SECTION_FONT
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))

    _autosize_columns(ws)

    # Summary sheet
    ws_meta = wb.create_sheet("About")
    ws_meta.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    ws_meta.append(["Records", len(eligible)])
    ws_meta.append([
        "Filter",
        "status=active, account_type=domain, has office phone OR cell OR extension",
    ])
    _autosize_columns(ws_meta)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
